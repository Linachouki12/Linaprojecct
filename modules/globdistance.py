import os
from math import radians, sin, cos, acos
import pandas as pd
import math
import glob
import time
import openpyxl
from openpyxl.styles import PatternFill


RED_COLOR_HEX = "FFFFC7CE"
red_fill = PatternFill(start_color=RED_COLOR_HEX, end_color=RED_COLOR_HEX, fill_type="solid")

def angle_diff_deg(az1, az2):
    """Plus petite différence d'angle entre deux azimuts """
    diff = abs(az1 - az2) % 360
    return diff if diff <= 180 else 360 - diff

def safe_float(val):
    try:
        return float(str(val).replace(',', '.'))
    except:
        return None
    
"""
def supprimer_fichiers_distance(base_folder, site_names):
    distance_folder = os.path.join(base_folder, "LTE_Distances")
    for site in site_names:
        pattern = os.path.join(distance_folder, site, "*Distance*.xlsx")
        for file_path in glob.glob(pattern):
            try:
                # Attendre si fichier utilisé
                for _ in range(3):
                    try:
                        # Tentative de suppression
                        os.remove(file_path)
                        print(f" Fichier supprimé : {file_path}")
                        break
                    except PermissionError as pe:
                        print(f" Fichier en cours d'utilisation (tentative de suppression reportée) : {file_path}")
                        time.sleep(1)  # attendre 1 seconde avant nouvelle tentative
                    except Exception as e:
                        print(f"Erreur lors de la suppression de {file_path} : {e}")
                        break
            except Exception as e:
                print(f"Problème inattendu avec {file_path} : {e}")   

"""

def bearing_from_to_deg(lat1, lon1, lat2, lon2):
    """Retourne l'azimut (angle) en degrés du point (lat1, lon1) vers (lat2, lon2)"""
    lat1 = math.radians(float(lat1))
    lat2 = math.radians(float(lat2))
    diff_long = math.radians(float(lon2) - float(lon1))
    x = math.sin(diff_long) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - (
        math.sin(lat1) * math.cos(lat2) * math.cos(diff_long)
    )
    initial_bearing = math.atan2(x, y)
    return (math.degrees(initial_bearing) + 360) % 360

def bearing_category(bearing_deg):
    if 0 <= bearing_deg < 90:
        return '0-90'
    elif 90 <= bearing_deg < 180:
        return '90-180'
    elif 180 <= bearing_deg < 270:
        return '180-270'
    elif 270 <= bearing_deg < 360:
        return '270-360'
    else:
        return 'invalide'

def custom_distance(lon1, lat1, lon2, lat2):
    """Calcule la distance en km entre deux points GPS"""
    lat1_rad = radians(lat1)
    lat2_rad = radians(lat2)
    lon_diff_rad = radians(lon1 - lon2)
    
    distance = acos(
        sin(lat1_rad) * sin(lat2_rad) + 
        cos(lat1_rad) * cos(lat2_rad) * cos(lon_diff_rad)
    ) * 6371  # Rayon de la Terre en km
    return distance
def calculate_distance(base_folder, site_names):
    print(f"Liste des sites disponibles : {site_names}")

    cell_plan_file = os.path.join(base_folder, "LTE_cell_Plan_HU_MasterFile.xlsx")
    ho_file = os.path.join(base_folder, 'HO_BAR_3G3G.xlsx')
    distance_folder = os.path.join(base_folder, "LTE_Distances")
    os.makedirs(distance_folder, exist_ok=True)

    try:
        # Lire uniquement les colonnes utiles
        df_cp = pd.read_excel(cell_plan_file, usecols=['CellName', 'Longitude_Sector', 'Latitude_Sector', 'Azimuth'])
        df_ho = pd.read_excel(ho_file)

        df_cp.columns = df_cp.columns.str.strip()
        df_ho.columns = df_ho.columns.str.strip()

        df_cp['Azimuth'] = pd.to_numeric(
            df_cp['Azimuth'].astype(str).str.replace('°', '', regex=False).str.strip(),
            errors='coerce'
        ).fillna(0)

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier Excel: {e}")
        return

    for site in set(site_names):
        df_ho['eNodeB_Short'] = df_ho['eNodeB Name'].str[:8]
        df_site_ho = df_ho[df_ho['eNodeB_Short'] == site]
    

        if df_site_ho.empty:
            print(f"Aucune ligne HO trouvée pour {site}")
            continue

        # Supprimer la colonne temporaire
        df_site_ho = df_site_ho.drop(columns=['eNodeB_Short'])

        # Fusion coordonnées pour Local cell name
        df_site_ho = df_site_ho.merge(
            df_cp.rename(columns={
                'CellName': 'Local cell name',
                'Longitude': 'Longitude_Local',
                'Latitude': 'Latitude_Local',
                'Azimuth': 'Azimuth_Local'
            }),
            on='Local cell name',
            how='left'
        )

        # Fusion coordonnées pour Target Cell Name
        df_site_ho = df_site_ho.merge(
            df_cp.rename(columns={
                'CellName': 'Target Cell Name',
                'Longitude': 'Longitude_Target',
                'Latitude': 'Latitude_Target',
                'Azimuth': 'Azimuth_Target'
            }),
            on='Target Cell Name',
            how='left'
        )
        
        df_site_ho.dropna(subset=[
                'Target Cell Name'
            ], inplace=True)
        df_site_ho['Target eNodeB ID']=df_site_ho['Target Cell Name'].str[:8]
        required_columns = [
                'Date',
                'eNodeB Name',
                'Local cell name',
                'Longitude_Sector_x',
                'Latitude_Sector_x',
                'Azimuth_Local',
                'Target eNodeB ID',
                'Target Cell Name',
                'Longitude_Sector_y',
                'Latitude_Sector_y',
                'Azimuth_Target',
                'L.HHO.NCell.ExecAttOut'
                
            ]
        df_site_ho = df_site_ho[required_columns]

        site_dir = os.path.join(distance_folder, site)
        os.makedirs(site_dir, exist_ok=True)
        


        source_cells = df_cp[df_cp['CellName'].str[:8] == site]
        if source_cells.empty:
            print(f"Aucune cellule trouvée pour le site {site} dans le cell plan.")
            continue



        distances = []
        angle_diffs = []
        angle_valids = []
        for idx, row in df_site_ho.iterrows():
            lon_src = safe_float(row['Longitude_Sector_x'])
            lat_src = safe_float(row['Latitude_Sector_x'])
            lon_tgt = safe_float(row['Longitude_Sector_y'])
            lat_tgt = safe_float(row['Latitude_Sector_y'])
            az_target = row['Azimuth_Target']
        
            if None in (lon_src, lat_src, lon_tgt, lat_tgt) or pd.isna(az_target):
                    distances.append(None)
                    angle_diffs.append(None)
                    angle_valids.append(None)
                    continue
            
            distance=custom_distance(lon_src, lat_src, lon_tgt, lat_tgt)
            az= bearing_from_to_deg(lat_src, lon_src, lat_tgt, lon_tgt)
            diff_angle = angle_diff_deg(az_target, az)
        

            distances.append(round(distance, 3))
            angle_diffs.append(round(diff_angle, 2))
            angle_valids.append(diff_angle)
        

        df_site_ho['Distance_km'] = distances
        df_site_ho['Angle_diff_deg'] = angle_diffs
        df_site_ho['Angle_valid'] = angle_valids
        
        dist_output_cols = [
            'Date',
            'eNodeB Name',
            'Local cell name',
            'Longitude_Sector_x',
            'Latitude_Sector_x',
            'Azimuth_Local',
            'Target Cell Name',
            'Longitude_Sector_y',
            'Latitude_Sector_y',
            'Azimuth_Target',
            'Distance_km',
            'Angle_diff_deg',
            'Angle_valid',
            'L.HHO.NCell.ExecAttOut'
        ]
        df_site_ho.sort_values(by="Distance_km", ascending=True, inplace=True)
        df_site_ho.drop_duplicates(subset=[
                'eNodeB Name', 'Longitude_Sector_x', 'Latitude_Sector_x', 'Azimuth_Local',
                'Target Cell Name', 'Longitude_Sector_y', 'Latitude_Sector_y', 'Azimuth_Target'
            ], inplace=True)
        att_sum_by_local_cell = df_site_ho.groupby("Local cell name")["L.HHO.NCell.ExecAttOut"].transform("sum")
        #print(f"la somme est égale à : {att_sum_by_local_cell}")
        df_site_ho["HO_attout_Percent_per_cell"] = df_site_ho["L.HHO.NCell.ExecAttOut"] / att_sum_by_local_cell * 100
        
# Affecter une valeur très élevée à ces cellules (pour créer des overshoots simulés)
        dist_output_cols.append("HO_attout_Percent_per_cell")


        overshoot_flags=[]
        for idx, row in df_site_ho.iterrows():
            cell = row['Local cell name']
            dist = row['Distance_km']
            percent = row.get('HO_attout_Percent_per_cell')
            if dist <= 3 or pd.isna(percent):
                overshoot_flags.append(False)
                continue

            subset_near = df_site_ho[
                (df_site_ho['Local cell name'] == cell) &
                (df_site_ho['Distance_km'] <= 3) &
                (~df_site_ho['HO_attout_Percent_per_cell'].isna())
            ]
            if subset_near.empty:
                overshoot_flags.append(False)
                continue

            avg_percent_near = subset_near['HO_attout_Percent_per_cell'].mean()

            if percent > avg_percent_near * 1.3:
                overshoot_flags.append(True)
            else:
                overshoot_flags.append(False)
        
        df_site_ho['Overshoot_by_HO_Percent'] = overshoot_flags
        distance_output_file = os.path.join(site_dir, f"{site}_Distance.xlsx")
        
        if 'Overshoot_by_HO_Percent' not in dist_output_cols:
            dist_output_cols.append("Overshoot_by_HO_Percent")
        df_site_ho[dist_output_cols].to_excel(distance_output_file, index=False)
        print(f" Fichier Distance enregistré : {distance_output_file}")

        wb = openpyxl.load_workbook(distance_output_file)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        col_overshoot = header.index("Overshoot_by_HO_Percent") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_overshoot)
            if cell.value is True:
        # Colorier toute la ligne en rouge
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill
            else:
                # Pas de couleur (par défaut)
                pass

        wb.save(distance_output_file)
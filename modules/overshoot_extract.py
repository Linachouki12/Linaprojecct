import os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


RED_COLOR_HEX = "FFFFC7CE"
red_fill = PatternFill(start_color=RED_COLOR_HEX, end_color=RED_COLOR_HEX, fill_type="solid")


def is_red(cell):
    fill = cell.fill
    if fill and fill.fgColor is not None:
        if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
            return fill.fgColor.rgb.upper() == RED_COLOR_HEX
    return False


def extract_overshoot_cells(base_folder, site_names):
    overshoot_cells = {}
    for site in site_names:
        file_path = os.path.join(base_folder, "LTE_Distances", site, f"{site}_Distance.xlsx")
        if not os.path.exists(file_path):
            continue
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        # Trouver les colonnes Source Cell Name et Target Cell Name
        header = [cell.value for cell in ws[1]]
        try:
            source_idx = header.index("Local cell name") + 1
            target_idx = header.index("Target Cell Name") + 1
            date_idx = header.index('Date')+1
            
        except ValueError:
            wb.close()
            continue

        overshoot_cells[site] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            source_cell = row[source_idx-1]
            cell_date= row[date_idx-1]
            target_cell = row[target_idx-1]
            # Vérifier si la couleur de fond est rouge (RGB:FFC7CE)
            
            if is_red(source_cell):
                overshoot_cells[site].append({
                    "Date" : cell_date.value,
                    "source": source_cell.value,
                    "target": target_cell.value
                })
    wb.close()
    return overshoot_cells 